import flet as ft
from openpyxl import load_workbook, Workbook
import os

# Fungsi untuk membaca data dari file Excel
def load_data_from_excel(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Mengabaikan header
        data.append(row)
    return data

# Fungsi untuk menambah data ke file Excel
def add_data_to_excel(file_name, name, age, subscription, employment):
    if not os.path.exists(file_name):
        # Jika file tidak ada, buat dengan struktur yang diperlukan
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Name", "Age", "Subscription", "Employment"])  # Menambah header
        workbook.save(file_name)

    # Coba buka workbook untuk read/write
    try:
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        # Hitung ID baru berdasarkan jumlah baris
        next_id = sheet.max_row  # ID adalah nomor baris berikutnya (karena header ada di baris pertama)

        # Tambahkan row baru ke sheet
        sheet.append([next_id, name, age, subscription, employment])

        # Simpan workbook dan pastikan file ditutup
        workbook.save(file_name)
        workbook.close()  # Pastikan workbook ditutup

    except PermissionError:
        # Tangani error jika terjadi masalah dengan izin file
        print(f"PermissionError: Unable to write to file '{file_name}'. Please check the file permissions.")

def main(page: ft.Page):
    page.title = "User Subscription Table"
    page.theme_mode = ft.ThemeMode.LIGHT

    # Nama file Excel (gunakan path absolut untuk menghindari masalah)
    excel_file = r"D:\Kuliah\SEMESTER 7\Mobile Application Development\Project\mobile_praktik_1\data\user_data.xlsx"

    # Pastikan folder 'data' ada, jika tidak buat foldernya
    if not os.path.exists("data"):
        os.makedirs("data")

    # Inisialisasi file Excel jika belum ada
    try:
        load_workbook(excel_file)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Name", "Age", "Subscription", "Employment"])  # Header
        workbook.save(excel_file)

    # Definisikan DataTable
    data_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("ID")),
            ft.DataColumn(ft.Text("Name")),
            ft.DataColumn(ft.Text("Age")),
            ft.DataColumn(ft.Text("Subscription")),
            ft.DataColumn(ft.Text("Employment")),
        ],
        rows=[],  # Inisialisasi rows kosong
    )

    # Fungsi untuk menampilkan data dari Excel ke dalam DataTable
    def tampil_user():
        data_table.rows.clear()  # Kosongkan data sebelum diisi
        data = load_data_from_excel(excel_file)
        if not data:  # Cek apakah data kosong
            page.snack_bar = ft.SnackBar(ft.Text("No data found in Excel!"))
            page.snack_bar.open = True
        for row in data:
            data_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(row[0]))),  # ID
                        ft.DataCell(ft.Text(row[1])),  # Name
                        ft.DataCell(ft.Text(str(row[2]))),  # Age
                        ft.DataCell(ft.Text(row[3])),  # Subscription
                        ft.DataCell(ft.Text(row[4])),  # Employment
                    ]
                )
            )
        page.update()

    # Buat form input untuk menambah data baru
    name_input = ft.TextField(label="Name")
    age_input = ft.TextField(label="Age")
    subscription_input = ft.Dropdown(
        label="Subscription",
        options=[
            ft.dropdown.Option("Subscribed"),
            ft.dropdown.Option("Not Subscribed"),
            ft.dropdown.Option("Other"),
        ],
    )
    employed_input = ft.Checkbox(label="Employed")

    # Aksi ketika menekan tombol Insert untuk menambah row baru
    def add_row(e):
        # Pastikan nilai Age adalah angka
        if age_input.value.isdigit():
            try:
                # Tambah data ke Excel
                add_data_to_excel(
                    excel_file, 
                    name_input.value, 
                    age_input.value, 
                    subscription_input.value, 
                    "Employed" if employed_input.value else "Unemployed"
                )

                # Bersihkan input setelah menambah data
                name_input.value = ""
                age_input.value = ""
                subscription_input.value = None
                employed_input.value = False

                # Refresh DataTable setelah menambah data baru
                tampil_user()

            except PermissionError:
                # Tampilkan pesan error jika gagal menulis ke file
                page.snack_bar = ft.SnackBar(ft.Text("Permission denied: Unable to write to the file!"))
                page.snack_bar.open = True
                page.update()
        else:
            # Tampilkan pesan error jika Age bukan angka
            page.snack_bar = ft.SnackBar(ft.Text("Age must be a number!"))
            page.snack_bar.open = True
            page.update()

    # Layout
    page.add(
        ft.Row([
            ft.Container(
                ft.Column([
                    ft.Text("Insert Row", size=18, weight="bold"),
                    name_input,
                    age_input,
                    subscription_input,
                    employed_input,
                    ft.ElevatedButton("Insert", on_click=add_row, color=ft.colors.BLUE),
                    ft.Switch(label="Mode", value=True),
                ], expand=1),
                border=ft.border.all(1, ft.colors.GREY),  # Tambahkan border abu-abu
                padding=10,
                height=400,
            ),
            ft.Container(
                ft.Column([
                    ft.VerticalDivider(),
                    data_table,
                ]),
                border=ft.border.all(1, ft.colors.GREY),  # Tambahkan border abu-abu
                padding=10,
                height=400,
            )
        ])
    )

    # Panggil fungsi untuk menampilkan data dari Excel
    tampil_user()

# Jalankan aplikasi
ft.app(target=main)
